"""
Case: 爬取https://www.redecor.com/全部的challenge详情 -- 三级页面文字
Author: yangwen
Time: 2020/12/11 11:21 上午
"""
import requests
import os
import re
import json
import openpyxl

from lxml import etree
from openpyxl import load_workbook
from datetime import datetime

'''
| Title | User names | Prices | Products used in this design |
| ----- | ---------- | ------ | ---------------------------- |
|       |            |        |                              |
'''


def get_data(fig_id):
    # 存储获取的数据到data_list
    data_list = []

    # 获取该fig_id的challenge详情url
    fig_detail_url = 'https://reworks.fi/challenge/' + fig_id
    # 获取该fig_id的challenge详情页数据
    fig_detail_text = requests.get(url=fig_detail_url, headers=headers).text
    fig_detail_tree = etree.HTML(fig_detail_text)
    # 获取count #
    fig_column_list = fig_detail_tree.xpath('//div[@class="columns is-multiline is-mobile"]/div')
    if not fig_column_list:
        print('图片地址解析错误！！！ Try next...')
        return

    # 遍历三级页面的所有column，获取user names、products used in this design, 存储到user_list
    counter = 0
    user_list = []
    for column in fig_column_list:
        if counter >= limit:
            break
        else:
            counter += 1
            # print('正在爬取', title, '第%d个用户数据...' % counter)
        # 获取每个column的user name #
        fig_user_name = column.xpath('.//h6/text()')[0]
        # 获取该user name的三级页面的url
        fig_user_name_url = 'https://reworks.fi' + column.xpath('.//a/@href')[0]
        # 获取该user name的三级页面数据
        fig_user_text = requests.get(url=fig_user_name_url, headers=headers).text
        fig_user_tree = etree.HTML(fig_user_text)
        fig_products_div = fig_user_tree.xpath('//div[@class="columns is-multiline is-mobile"]/div')
        # 获取price #
        fig_price_div = fig_user_tree.xpath('//div[@class="level-right level-divided is-mobile"]/div')[0]
        fig_price = fig_price_div.xpath('./text()')[0]
        # 获取products used in this design, 存储到列表里 #
        fig_products_list, user_temp_list = [], []
        for product_div in fig_products_div:
            product = product_div.xpath('.//img/@alt')[0]
            fig_products_list.append(product)

        # 存储
        user_temp_list.append(fig_user_name)
        user_temp_list.append(fig_price)
        user_temp_list.append(fig_products_list)
        user_list.append(user_temp_list)
    data_list.extend(user_list)
    return data_list


def generate_excel(detail_head):
    wb = openpyxl.Workbook()
    # 创建sheet并设置列标题
    ws = wb.worksheets[0]
    ws.title = 'detail'
    ws.append(detail_head)

    wb.save('redecor_excel_detail' + '.xlsx')

    # # 获取当前日期，得到一个datetime对象如：(2016, 8, 9, 23, 12, 23, 424000)
    # today = datetime.today()
    # # 将获取到的datetime对象仅取日期如：2016-8-9
    # today_date = datetime.date(today)
    # # 以name+当前日期作为excel名称保存。
    # wb.save('redecor_excel_detail-' + str(today_date) + '.xlsx')


def write_data_to_excel(user_row, detail_title_row, data_list):
    # 读取excel
    wb = load_workbook('redecor_excel_detail.xlsx')
    ws_detail = wb.worksheets[0]

    ws_detail_title = data_list.pop(0)
    detail_list_len = len(data_list)
    # 写入数据到ws_detail
    ws_detail.cell(row=detail_title_row, column=1, value=ws_detail_title)
    for j in range(detail_list_len):
        for k in range(len(data_list[0])):
            value = data_list[j]
            if k != detail_length - 2:
                ws_detail.cell(row=user_row, column=2 + k, value=value[k])
            else:
                ws_detail.cell(row=user_row, column=2 + k, value=','.join(value[k]))
        user_row += 1
    detail_title_row += detail_list_len
    wb.save('redecor_excel_detail.xlsx')
    print(ws_detail_title, '数据写入成功！！！')
    return user_row, detail_title_row


def get_dyna_challenge(fig_id, user_row, detail_title_row):
    try:
        while True:
            post_url = 'https://api.redecor.com/graphql'
            # 请求新图片依赖参数: "after": fig_id (最后一张图片id)
            payload = {"operationName": "ALL_CHALLENGES_QUERY",
                       "variables": {"type": "closed", "after": fig_id, "limit": 18},
                       "query": "query ALL_CHALLENGES_QUERY($type: ChallengeState = all, "
                                "$after: ID = null, $limit: Int = 18) "
                                "{\n  listChallenges(type: $type, after: $after, limit: $limit) {\n    id\n    title\n "
                                "dateEnd\n    thumb\n    designs(limit: 1, orderBy: rating) {\n      thumb\n      "
                                "__typename\n    }\n    __typename\n  }\n}\n"}

            fig_json = requests.post(url=post_url, headers=headers, data=json.dumps(payload)).json()
            challenges_list = fig_json['data']['listChallenges']
            # 获取最后一个fig_id， 作为下一次请求的after参数
            fig_id = challenges_list[-1]['id']
            print('fig_id:', fig_id)
            for challenges in challenges_list:
                # 获取请求二级页面所需的"id"参数值
                new_fig_title = challenges['title']
                new_fig_id = challenges['id']
                print('new_fig_id:', new_fig_id)
                new_data_list = get_data(new_fig_id)
                if not new_data_list:
                    continue
                else:
                    new_data_list.insert(0, new_fig_title)
                # 将数据写入表格
                user_row, detail_title_row = write_data_to_excel(user_row, detail_title_row, new_data_list)
            print("------------------Load more Challenges------------------")
    except:
        return fig_id, user_row, detail_title_row


if __name__ == '__main__':
    # 请求头
    headers = {
        'content-type': 'application/json; charset=UTF-8',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/87.0.4280.67 Safari/537.36 '
    }
    # 表格列标题
    detail_head = ['Title', 'User names', 'Prices', 'Products used in this design']
    # 根据列标题生成excel
    if not os.path.exists('./redecor_excel_detail.xlsx'):
        generate_excel(detail_head)
    # 每个challenge里爬取图片数量限制
    limit = 10
    # 表格相关
    user_row, detail_title_row = 8349, 8349
    detail_length = len(detail_head)
    print('======================开始爬取======================')

    # # 爬取首页图片 #
    # url = 'https://reworks.fi/'
    # # 获取首页数据
    # page_text = requests.get(url=url, headers=headers).text
    # # 使用xpath进行数据解析
    # tree = etree.HTML(page_text)
    # div_list = tree.xpath('//div[@class="columns is-multiline"]/div')
    # # 遍历页面，获取各项数据
    # for div in div_list:
    #     # 获取title #
    #     title = div.xpath('./figure/@title')[0]
    #     # 获取fig_id
    #     a_href = div.xpath('.//a/@href')[0]
    #     fig_id = re.findall(r'\d+', a_href)[0]
    #     # 获取各项数据
    #     data_list = get_data(fig_id)
    #     if not data_list:
    #         continue
    #     else:
    #         data_list.insert(0, title)
    #     # 将数据写入表格
    #     user_row, detail_title_row = write_data_to_excel(user_row, detail_title_row, data_list)
    print("------------------首页图片爬取结束------------------")

    # 爬取动态加载图片 #
    # 获取第一个fig_id
    # a_href = div_list[-1].xpath('.//a/@href')[0]
    # fig_id = re.findall(r'\d+', a_href)[0]
    last_fig_id = '0'
    fig_id = '2301'

    # 获取动态数据
    try:
        last_fig_id, user_row, detail_title_row = get_dyna_challenge(fig_id, user_row, detail_title_row)
    except:
        # 断线重连
        # for _ in range(1, 1000):
        #     last_fig_id, user_row, detail_title_row = get_dyna_challenge(last_fig_id, user_row, detail_title_row)
        #     print('last_fig_id:', last_fig_id)
        #     print("********************网络异常！！！第%d次尝试重新连接********************" % _)
        print('======================爬取结束======================')